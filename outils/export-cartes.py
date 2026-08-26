# -*- coding: utf-8 -*-
"""
Exporte la banque de cartes (Question / Defi / Cap ou pas) vers Cartes.xlsx.

    python outils/export-cartes.py

Un onglet par categorie. Les cartes qui contiennent encore le marqueur
« [ ... complete ici ] » sont signalees « A COMPLETER » sur fond orange :
c'est exactement ce qu'il reste a relire et a remplir.

Le classeur est volontairement exclu du depot (.gitignore).
"""
import io, os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHEMIN_JS = os.path.join(RACINE, "cards-data.js")
CHEMIN_XLSX = os.path.join(RACINE, "Cartes.xlsx")
MARQUEUR = u"[ \u2026 compl\u00e8te ici ]"

LIGNE = re.compile(r"^\s*\{ id:'([^']+)', cat:'([^']+)', text:\"(.*)\", pts:(\d+)(, photo:true)? \},\s*$")

cartes = []
for ligne in io.open(CHEMIN_JS, encoding="utf-8").read().split("\n"):
    m = LIGNE.match(ligne)
    if m:
        cartes.append({"id": m.group(1), "cat": m.group(2),
                       "text": m.group(3), "pts": int(m.group(4)),
                       "photo": bool(m.group(5))})

ONGLETS = [("gage", "Cap ou pas"), ("question", "Question"), ("defi", "Defi")]

POLICE = "Arial"
titre_fill = PatternFill("solid", fgColor="2B1B2E")
titre_font = Font(name=POLICE, bold=True, color="FFFFFF", size=11)
modifiable = PatternFill("solid", fgColor="FFF7E0")
a_completer = PatternFill("solid", fgColor="FFD9A0")
trait = Side(style="thin", color="D9D9D9")
bord = Border(left=trait, right=trait, top=trait, bottom=trait)

wb = openpyxl.Workbook()
wb.remove(wb.active)

for cat, nom in ONGLETS:
    lot = [x for x in cartes if x["cat"] == cat]
    reste = len([x for x in lot if MARQUEUR in x["text"]])
    ws = wb.create_sheet(nom)

    ws["A1"] = nom + " (" + str(len(lot)) + " cartes, " + (
        str(reste) + " a completer)" if reste else "toutes completes)")
    ws["A1"].font = Font(name=POLICE, bold=True, size=13)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    ws["A2"] = (u"Remplace le marqueur « " + MARQUEUR + u" » par le texte voulu, directement dans "
                u"la colonne Texte. Modifie librement les cellules creme. Colonne Action : SUPPRIMER "
                u"pour retirer une carte. Pour en ajouter, remplis une nouvelle ligne en laissant "
                u"l'id vide. Colonne Photo : OUI = la reponse a cette carte est une photo, "
                u"affichee 3 secondes chez l'autre. Ne modifie jamais un id existant ni l'onglet "
                u"d'une carte : c'est l'id qui porte le masquage et les modifications faites "
                u"depuis l'appli.")
    ws["A2"].font = Font(name=POLICE, italic=True, size=9, color="7F7F7F")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 42

    for c, titre in enumerate(["id", "Etat", "Texte", "Points", "Photo", "Action"], 1):
        cell = ws.cell(row=4, column=c, value=titre)
        cell.font = titre_font
        cell.fill = titre_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bord

    for i, x in enumerate(lot):
        r = 5 + i
        idc = ws.cell(row=r, column=1, value=x["id"])
        idc.font = Font(name=POLICE, size=9, color="808080")
        idc.alignment = Alignment(horizontal="center", vertical="top")

        vide = MARQUEUR in x["text"]
        ec = ws.cell(row=r, column=2, value=("A COMPLETER" if vide else "OK"))
        ec.alignment = Alignment(horizontal="center", vertical="top")
        if vide:
            ec.font = Font(name=POLICE, size=9, bold=True, color="9C5700")
            ec.fill = a_completer
        else:
            ec.font = Font(name=POLICE, size=9, color="217346")

        ws.cell(row=r, column=3, value=x["text"])
        ws.cell(row=r, column=4, value=x["pts"]).alignment = Alignment(horizontal="center", vertical="top")
        ph = ws.cell(row=r, column=5, value=("OUI" if x["photo"] else ""))
        ph.alignment = Alignment(horizontal="center", vertical="top")
        ph.font = Font(name=POLICE, size=9, bold=True, color="C0392B")
        for c in (3, 6):
            ws.cell(row=r, column=c).alignment = Alignment(wrap_text=True, vertical="top")
        for c in range(3, 7):
            ws.cell(row=r, column=c).fill = modifiable
            if c != 5:
                ws.cell(row=r, column=c).font = Font(name=POLICE, size=10)
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = bord

    for c, largeur in enumerate([8, 13, 86, 8, 8, 13], 1):
        ws.column_dimensions[get_column_letter(c)].width = largeur
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:F" + str(4 + len(lot))

wb.save(CHEMIN_XLSX)
print("classeur ecrit :", CHEMIN_XLSX)
for cat, nom in ONGLETS:
    lot = [x for x in cartes if x["cat"] == cat]
    print("  ", nom, ":", len(lot), "cartes,",
          len([x for x in lot if MARQUEUR in x["text"]]), "a completer")
