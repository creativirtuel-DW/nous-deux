# -*- coding: utf-8 -*-
"""
Exporte la banque de questions Battle vers Questions-Battle.xlsx.

    python outils/export-questions.py

Un onglet par catégorie. Les questions déjà relues portent leur date de
validation en vert ; celles jamais relues sont signalées « À VALIDER » sur
fond orange, pour qu'on voie d'un coup d'œil ce qui reste à passer en revue.

Le classeur est volontairement exclu du dépôt (.gitignore).
"""
import io, os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHEMIN_JS = os.path.join(RACINE, "battle-data.js")
CHEMIN_XLSX = os.path.join(RACINE, "Questions-Battle.xlsx")

s = io.open(CHEMIN_JS, encoding="utf-8").read()

questions = []
for ligne in s.split("\n"):
    if "cat:'" not in ligne or "id:'" not in ligne:
        continue
    valide = re.search(r"v:'([^']+)'", ligne)
    questions.append({
        "id": re.search(r"id:'([^']+)'", ligne).group(1),
        "cat": re.search(r"cat:'([^']+)'", ligne).group(1),
        "q": re.search(r'q:"([^"]*)"', ligne).group(1),
        "a": re.findall(r'"([^"]*)"', ligne.split("a:[", 1)[1])[:4],
        "v": valide.group(1) if valide else "",
    })

POLICE = "Arial"
titre_fill = PatternFill("solid", fgColor="2B1B2E")
titre_font = Font(name=POLICE, bold=True, color="FFFFFF", size=11)
modifiable = PatternFill("solid", fgColor="FFF7E0")
a_valider = PatternFill("solid", fgColor="FFD9A0")
trait = Side(style="thin", color="D9D9D9")
bord = Border(left=trait, right=trait, top=trait, bottom=trait)

wb = openpyxl.Workbook()
wb.remove(wb.active)

entetes = ["id", "Valide le", "Question", "Reponse 1", "Reponse 2",
           "Reponse 3", "Reponse 4", "Action", "Remarque"]
onglets = [("Intime", "coquin"), ("Quotidien", "quotidien"), ("Souvenir", "souvenir")]

for nom_onglet, cat in onglets:
    lot = [x for x in questions if x["cat"] == cat]
    reste = len([x for x in lot if not x["v"]])
    ws = wb.create_sheet(nom_onglet)

    ws["A1"] = ("Questions Battle - " + nom_onglet + " (" + str(len(lot)) + " questions, "
                + (str(reste) + " a valider)" if reste else "toutes validees)"))
    ws["A1"].font = Font(name=POLICE, bold=True, size=13)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=9)
    ws["A2"] = ("Colonne 'Valide le' : date de ta derniere relecture. Les lignes orange n'ont jamais ete relues. "
                "Modifie librement les cellules creme. Colonne Action : SUPPRIMER pour retirer une question. "
                "Pour en ajouter, remplis une nouvelle ligne avec un id neuf. Ne modifie jamais un id existant.")
    ws["A2"].font = Font(name=POLICE, italic=True, size=9, color="7F7F7F")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 34

    for c, titre in enumerate(entetes, 1):
        cell = ws.cell(row=4, column=c, value=titre)
        cell.font = titre_font
        cell.fill = titre_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = bord

    for i, x in enumerate(lot):
        r = 5 + i
        idc = ws.cell(row=r, column=1, value=x["id"])
        idc.font = Font(name=POLICE, size=9, color="808080")
        idc.alignment = Alignment(horizontal="center", vertical="top")

        vc = ws.cell(row=r, column=2, value=(x["v"] if x["v"] else "A VALIDER"))
        vc.alignment = Alignment(horizontal="center", vertical="top")
        if x["v"]:
            vc.font = Font(name=POLICE, size=9, color="217346")
        else:
            vc.font = Font(name=POLICE, size=9, bold=True, color="9C5700")
            vc.fill = a_valider

        ws.cell(row=r, column=3, value=x["q"])
        for j in range(4):
            ws.cell(row=r, column=4 + j, value=x["a"][j])

        for c in range(3, 10):
            cel = ws.cell(row=r, column=c)
            cel.font = Font(name=POLICE, size=10)
            cel.alignment = Alignment(wrap_text=True, vertical="top")
            cel.fill = modifiable
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = bord

    for c, largeur in enumerate([8, 12, 48, 23, 23, 23, 23, 12, 26], 1):
        ws.column_dimensions[get_column_letter(c)].width = largeur
    ws.freeze_panes = "A5"
    ws.auto_filter.ref = "A4:I" + str(4 + len(lot))

wb.save(CHEMIN_XLSX)
print("classeur ecrit :", CHEMIN_XLSX)
for nom_onglet, cat in onglets:
    lot = [x for x in questions if x["cat"] == cat]
    print("  ", nom_onglet, ":", len(lot), "questions,",
          len([x for x in lot if not x["v"]]), "a valider")
