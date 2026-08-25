# -*- coding: utf-8 -*-
"""
Exporte l'inventaire des cartes par niveau de points vers Cartes-fortes.xlsx.

    python outils/export-cartes-fortes.py

Onglet « Repartition » : combien de cartes existent pour chaque valeur de
points, par categorie, avec le palier de la Soiree guidee qui pioche dedans.
Onglet « Cartes fortes » : le texte de toutes les cartes a 40 points ou plus.

Sert a voir ou la banque est trop mince : un palier qui n'a que deux cartes
possibles ramene toujours les deux memes en fin de soiree.

Le classeur est volontairement exclu du depot (.gitignore).
"""
import io, os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHEMIN_JS = os.path.join(RACINE, "cards-data.js")
CHEMIN_SOIREE = os.path.join(RACINE, "soiree.js")
CHEMIN_XLSX = os.path.join(RACINE, "Cartes-fortes.xlsx")
SEUIL_FORT = 40

CATS = [("gage", "Cap ou pas"), ("question", "Question"), ("defi", "Defi")]
LIGNE = re.compile(r"^\s*\{ id:'([^']+)', cat:'([^']+)', text:\"(.*)\", pts:(\d+) \},\s*$")

cartes = []
for ligne in io.open(CHEMIN_JS, encoding="utf-8").read().split("\n"):
    m = LIGNE.match(ligne)
    if m:
        cartes.append({"id": m.group(1), "cat": m.group(2),
                       "text": m.group(3), "pts": int(m.group(4))})

# --- paliers de la Soiree guidee, relus directement dans soiree.js ---
niveaux = []
for m in re.finditer(r"id:'(\w+)',\s*emoji:'([^']*)',\s*label:'([^']+)',\s*max:(\d+),\s*paliers:\[([0-9, ]+)\]",
                     io.open(CHEMIN_SOIREE, encoding="utf-8").read()):
    niveaux.append({"id": m.group(1), "label": m.group(3), "max": int(m.group(4)),
                    "paliers": [int(x) for x in m.group(5).split(",")]})

POLICE = "Arial"
titre_fill = PatternFill("solid", fgColor="2B1B2E")
titre_font = Font(name=POLICE, bold=True, color="FFFFFF", size=11)
alerte = PatternFill("solid", fgColor="FFD9A0")
ok_fill = PatternFill("solid", fgColor="DDEFDD")
trait = Side(style="thin", color="D9D9D9")
bord = Border(left=trait, right=trait, top=trait, bottom=trait)

wb = openpyxl.Workbook()

# ================= ONGLET 1 : repartition =================
ws = wb.active
ws.title = "Repartition"
valeurs = sorted(set(c["pts"] for c in cartes))

ws["A1"] = "Combien de cartes pour chaque valeur de points (" + str(len(cartes)) + " cartes au total)"
ws["A1"].font = Font(name=POLICE, bold=True, size=13)
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws["A2"] = (u"La colonne « Fin de soiree » indique le niveau dont la DERNIERE carte se tire "
            u"dans cette valeur : c'est la qu'une banque trop mince se voit, le bouquet final "
            u"revenant toujours a l'identique. En orange, moins de 5 cartes disponibles.")
ws["A2"].font = Font(name=POLICE, italic=True, size=9, color="7F7F7F")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 40

entetes = ["Points", "Total"] + [nom for _, nom in CATS] + ["Fin de soiree"]
for c, t in enumerate(entetes, 1):
    cell = ws.cell(row=4, column=c, value=t)
    cell.font = titre_font
    cell.fill = titre_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bord

for i, v in enumerate(valeurs):
    r = 5 + i
    lot = [c for c in cartes if c["pts"] == v]
    ws.cell(row=r, column=1, value=v).font = Font(name=POLICE, bold=True, size=10)
    tot = ws.cell(row=r, column=2, value=len(lot))
    tot.font = Font(name=POLICE, bold=True, size=10)
    tot.fill = alerte if len(lot) < 5 else ok_fill
    for j, (cat, _nom) in enumerate(CATS):
        ws.cell(row=r, column=3 + j, value=len([c for c in lot if c["cat"] == cat]))
    fins = [n["label"] for n in niveaux if n["paliers"][-1] == v]
    ws.cell(row=r, column=3 + len(CATS), value=", ".join(fins) if fins else "")
    for c in range(1, 4 + len(CATS)):
        cel = ws.cell(row=r, column=c)
        cel.border = bord
        if cel.alignment.horizontal is None:
            cel.alignment = Alignment(horizontal="center")
        if not cel.font.bold:
            cel.font = Font(name=POLICE, size=10)

r = 5 + len(valeurs) + 1
ws.cell(row=r, column=1, value="Paliers de la Soiree guidee").font = Font(name=POLICE, bold=True, size=11)
for i, n in enumerate(niveaux):
    rr = r + 1 + i
    ws.cell(row=rr, column=1, value=n["label"]).font = Font(name=POLICE, bold=True, size=10)
    detail = " -> ".join(str(p) for p in n["paliers"])
    cel = ws.cell(row=rr, column=2, value=detail)
    cel.font = Font(name=POLICE, size=10)
    cel.alignment = Alignment(horizontal="left")

for c, l in enumerate([9, 9, 13, 12, 9, 18], 1):
    ws.column_dimensions[get_column_letter(c)].width = l
ws.freeze_panes = "A5"

# ================= ONGLET 2 : cartes fortes =================
ws2 = wb.create_sheet("Cartes fortes")
fortes = sorted([c for c in cartes if c["pts"] >= SEUIL_FORT],
                key=lambda c: (-c["pts"], c["cat"]))
ws2["A1"] = "Les " + str(len(fortes)) + " cartes a " + str(SEUIL_FORT) + " points ou plus"
ws2["A1"].font = Font(name=POLICE, bold=True, size=13)
ws2.merge_cells(start_row=2, start_column=1, end_row=2, end_column=4)
ws2["A2"] = (u"Ce sont elles qui ferment une soiree. Pour en ajouter, passe par Cartes.xlsx "
             u"(export-cartes.py / import-cartes.py) en mettant 100 dans la colonne Points.")
ws2["A2"].font = Font(name=POLICE, italic=True, size=9, color="7F7F7F")
ws2["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws2.row_dimensions[2].height = 26

for c, t in enumerate(["Points", "id", "Categorie", "Texte"], 1):
    cell = ws2.cell(row=4, column=c, value=t)
    cell.font = titre_font
    cell.fill = titre_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bord

nom_cat = dict(CATS)
for i, x in enumerate(fortes):
    r = 5 + i
    p = ws2.cell(row=r, column=1, value=x["pts"])
    p.font = Font(name=POLICE, bold=True, size=10, color="C0392B" if x["pts"] >= 100 else "9C5700")
    p.alignment = Alignment(horizontal="center", vertical="top")
    idc = ws2.cell(row=r, column=2, value=x["id"])
    idc.font = Font(name=POLICE, size=9, color="808080")
    idc.alignment = Alignment(horizontal="center", vertical="top")
    cc = ws2.cell(row=r, column=3, value=nom_cat.get(x["cat"], x["cat"]))
    cc.font = Font(name=POLICE, size=9)
    cc.alignment = Alignment(horizontal="center", vertical="top")
    tc = ws2.cell(row=r, column=4, value=x["text"])
    tc.font = Font(name=POLICE, size=10)
    tc.alignment = Alignment(wrap_text=True, vertical="top")
    for c in range(1, 5):
        ws2.cell(row=r, column=c).border = bord

for c, l in enumerate([9, 8, 13, 96], 1):
    ws2.column_dimensions[get_column_letter(c)].width = l
ws2.freeze_panes = "A5"
ws2.auto_filter.ref = "A4:D" + str(4 + len(fortes))

wb.save(CHEMIN_XLSX)
print("classeur ecrit :", CHEMIN_XLSX)
for v in sorted(valeurs, reverse=True)[:5]:
    print("  ", v, "points :", len([c for c in cartes if c["pts"] == v]), "cartes")
