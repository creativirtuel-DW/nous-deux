# -*- coding: utf-8 -*-
"""
Exporte la banque de gages aléatoires vers Gages-Photo.xlsx.

    python outils/export-gages.py

Un seul onglet. Les gages déjà relus portent leur date de validation en vert ;
ceux jamais relus sont signalés « À VALIDER » sur fond orange.

Le classeur est volontairement exclu du dépôt (.gitignore).
"""
import io, os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

RACINE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CHEMIN_JS = os.path.join(RACINE, "gages-data.js")
CHEMIN_XLSX = os.path.join(RACINE, "Gages-Photo.xlsx")

s = io.open(CHEMIN_JS, encoding="utf-8").read()

gages = []
for ligne in s.split("\n"):
    if "id:'" not in ligne or "type:'" not in ligne:
        continue
    gages.append({
        "id": re.search(r"id:'([^']+)'", ligne).group(1),
        "type": re.search(r"type:'([^']+)'", ligne).group(1),
        "v": re.search(r"v:'([^']*)'", ligne).group(1),
        "g": re.search(r'g:"([^"]*)"', ligne).group(1),
    })

POLICE = "Arial"
titre_fill = PatternFill("solid", fgColor="2B1B2E")
titre_font = Font(name=POLICE, bold=True, color="FFFFFF", size=11)
modifiable = PatternFill("solid", fgColor="FFF7E0")
a_valider = PatternFill("solid", fgColor="FFD9A0")
trait = Side(style="thin", color="D9D9D9")
bord = Border(left=trait, right=trait, top=trait, bottom=trait)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gages"

reste = len([x for x in gages if not x["v"]])
ws["A1"] = ("Gages aleatoires du Defi Photo (" + str(len(gages)) + " gages, "
            + (str(reste) + " a valider)" if reste else "tous valides)"))
ws["A1"].font = Font(name=POLICE, bold=True, size=13)

ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
ws["A2"] = ("Type : action (quelque chose a faire) ou verite (quelque chose a avouer). "
            "Modifie librement les cellules creme. Colonne Action : SUPPRIMER pour retirer un gage. "
            "Pour en ajouter, remplis une nouvelle ligne avec un id neuf (a### pour une action, v### pour une verite). "
            "Ne modifie jamais un id existant. Le texte s'adresse a celui qui realise le gage.")
ws["A2"].font = Font(name=POLICE, italic=True, size=9, color="7F7F7F")
ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[2].height = 42

entetes = ["id", "Type", "Valide le", "Gage", "Action", "Remarque"]
for c, titre in enumerate(entetes, 1):
    cell = ws.cell(row=4, column=c, value=titre)
    cell.font = titre_font
    cell.fill = titre_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = bord

for i, x in enumerate(gages):
    r = 5 + i
    idc = ws.cell(row=r, column=1, value=x["id"])
    idc.font = Font(name=POLICE, size=9, color="808080")
    idc.alignment = Alignment(horizontal="center", vertical="top")

    tc = ws.cell(row=r, column=2, value=x["type"])
    tc.alignment = Alignment(horizontal="center", vertical="top")
    tc.font = Font(name=POLICE, size=9, bold=True,
                   color="C0392B" if x["type"] == "action" else "2874A6")

    vc = ws.cell(row=r, column=3, value=(x["v"] if x["v"] else "A VALIDER"))
    vc.alignment = Alignment(horizontal="center", vertical="top")
    if x["v"]:
        vc.font = Font(name=POLICE, size=9, color="217346")
    else:
        vc.font = Font(name=POLICE, size=9, bold=True, color="9C5700")
        vc.fill = a_valider

    ws.cell(row=r, column=4, value=x["g"])
    for c in range(4, 7):
        cel = ws.cell(row=r, column=c)
        cel.font = Font(name=POLICE, size=10)
        cel.alignment = Alignment(wrap_text=True, vertical="top")
        cel.fill = modifiable
    for c in range(1, 7):
        ws.cell(row=r, column=c).border = bord

for c, largeur in enumerate([8, 10, 12, 76, 13, 28], 1):
    ws.column_dimensions[get_column_letter(c)].width = largeur
ws.freeze_panes = "A5"
ws.auto_filter.ref = "A4:F" + str(4 + len(gages))

wb.save(CHEMIN_XLSX)
print("classeur ecrit :", CHEMIN_XLSX)
print("  ", len(gages), "gages,", len([x for x in gages if x["type"] == "action"]), "actions,",
      len([x for x in gages if x["type"] == "verite"]), "verites,", reste, "a valider")
